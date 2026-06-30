import io
import re
import pandas as pd
import streamlit as st
from openpyxl.styles import PatternFill

# --- CONSTANTS ---
RATE_PER_LAKH = 435.0

# Define the exact required output columns in order based on the updated format file provided
OUTPUT_COLUMNS = [
    "Sr. no.",
    "Zone",
    "Branch Name",
    "Branch Code",
    "Loan Type",
    "Loan Account No.",
    "Name of Primary Loan borrower",
    "Loan Amount ",
    "Gender (Primary Loan )",
    "Date of Birth of Primary Loan borrower (DDMMMYYYY)",
    "Type of    Age Proof ",
    "Address (First Life)",
    "Address 1            (First Life)",
    "Address 2            (First Life)",
    "Pincode",
    "Mobile No",
    "Email Id",
    "Nominee Name",
    "Relationship of the Nominee with Insurance covered Person",
    "Nominee DOB(DDMMMYYYY)*please mention name of the month Eg 10Feb1991",
    "Nominee Gender",
    "Appointee Name",
    "Relationship with Borrower",
    "Appointee DOB(DDMMMYYYY)*please mention name of the month Eg 10Feb1991",
    "Loan Outstanding Amount",
    "Sum Assured",
    "Loan Disbursement Date (DDMMYYYY)",
    "Loan End date (DDMMYYYY)",
    " Loan Term (in months)",
    " Loan Term (Year)",
    "MAIN MEMBER AGE",
    "Rate",
    "Premium (Excl. GST)",
    "GST amount",
    "Total Premium (incl GST)",
]


# --- HELPER FUNCTIONS ---
def normalize_text(text: str) -> str:
    """Normalize headers for flexible matching by removing line breaks,

    extra spaces, and converting to lowercase.
    """
    if not isinstance(text, str):
        return ""
    text = text.replace("\n", " ").replace("\r", " ")
    text = re.sub(r"\s+", " ", text)
    return text.strip().lower()


def calculate_premium_metrics(sum_assured: float):
    """Calculate direct static values for Premium, GST, and Total Premium.

    Formula:
    Total Premium = (Sum Assured / 100,000) * 435
    Premium (Excl. GST) = Total Premium / 1.18
    GST amount = Total Premium - Premium (Excl. GST)
    """
    total_premium = (sum_assured / 100000.0) * RATE_PER_LAKH
    premium = total_premium / 1.18
    gst = total_premium - premium
    return round(premium, 2), round(gst, 2), round(total_premium, 2)


# --- APP SETUP ---
st.set_page_config(
    page_title="Group Term Life Insurance Premium Calculator", layout="wide"
)
st.title("Group Term Life (GTL) Insurance Premium Calculator")

# Sidebar navigation between modules
module = st.sidebar.radio("Select Module", ["Quick Calculator", "Bulk Calculator"])

# --- MODULE 1: QUICK PREMIUM CALCULATOR ---
if module == "Quick Calculator":
    st.header("Quick Premium Calculator")
    st.write("Calculate premium for an individual instantly.")

    opted_sa = st.number_input(
        "Sum Assured (₹)", min_value=0.0, step=50000.0, value=1000000.0
    )

    if opted_sa > 0:
        prem, gst_val, total_prem = calculate_premium_metrics(opted_sa)

        col1, col2, col3 = st.columns(3)
        col1.metric("Premium (Excl. GST)", f"₹{prem:,.2f}")
        col2.metric("GST Amount (18%)", f"₹{gst_val:,.2f}")
        col3.metric("Total Premium (Incl. GST)", f"₹{total_prem:,.2f}")
    else:
        st.warning("Please enter a Sum Assured greater than 0.")

# --- MODULE 2: BULK PREMIUM CALCULATOR ---
else:
    st.header("Bulk Premium Calculator")
    st.write(
        "Upload the client filled Excel file to generate strict matching premium records."
    )

    uploaded_file = st.file_uploader(
        "Upload Client Excel File (.xlsx)", type=["xlsx"]
    )

    if uploaded_file is not None:
        try:
            # Read input Excel file
            df_input = pd.read_excel(uploaded_file, engine="openpyxl")

            # Create header normalize maps
            norm_to_orig = {
                normalize_text(col): col for col in df_input.columns
            }

            # Map inputs strictly against structural layout
            target_name_norm = normalize_text("Name of Primary Loan borrower")
            target_sa_norm = normalize_text("Sum Assured")

            # Check presence of mandatory operational columns
            missing_cols = []
            if target_name_norm not in norm_to_orig:
                missing_cols.append("• Name of Primary Loan borrower")
            if target_sa_norm not in norm_to_orig:
                missing_cols.append("• Sum Assured")

            if missing_cols:
                st.error("Missing mandatory columns:\n" + "\n".join(missing_cols))
                st.stop()

            orig_name_col = norm_to_orig[target_name_norm]
            orig_sa_col = norm_to_orig[target_sa_norm]

            # Build rigid target matrix matching client specification template schema
            df_output = pd.DataFrame(columns=OUTPUT_COLUMNS)

            for col in OUTPUT_COLUMNS:
                norm_output_col = normalize_text(col)
                if norm_output_col in norm_to_orig:
                    df_output[col] = df_input[norm_to_orig[norm_output_col]]
                elif norm_output_col == target_name_norm:
                    df_output[col] = df_input[orig_name_col]
                elif norm_output_col == target_sa_norm:
                    df_output[col] = df_input[orig_sa_col]
                else:
                    df_output[col] = None

            # Enforce calculations natively based on row metrics
            df_output["Sum Assured"] = (
                pd.to_numeric(df_output["Sum Assured"], errors="coerce")
                .fillna(0)
                .astype(float)
            )

            premiums = []
            gsts = []
            total_premiums = []
            rates = []

            for sa in df_output["Sum Assured"]:
                if sa > 0:
                    p, g, tp = calculate_premium_metrics(sa)
                    premiums.append(p)
                    gsts.append(g)
                    total_premiums.append(tp)
                    rates.append(RATE_PER_LAKH)
                else:
                    premiums.append(0.0)
                    gsts.append(0.0)
                    total_premiums.append(0.0)
                    rates.append(RATE_PER_LAKH)

            # Assign properties back into the specific custom keys matching file headers
            df_output["Rate"] = rates
            df_output["Premium (Excl. GST)"] = premiums
            df_output["GST amount"] = gsts
            df_output["Total Premium (incl GST)"] = total_premiums

            # Fix standard positional counter index
            df_output["Sr. no."] = range(1, len(df_output) + 1)

            # Ensure schema configuration maps perfectly
            df_output = df_output[OUTPUT_COLUMNS]

            # --- METRICS DISPLAY SUMMARY PANEL ---
            st.subheader("Portfolio Summary")
            sum_col1, sum_col2, sum_col3 = st.columns(3)

            total_members = len(df_output)
            total_sa = df_output["Sum Assured"].sum()
            total_prem_sum = df_output["Total Premium (incl GST)"].sum()

            sum_col1.metric("Total Members", f"{total_members:,}")
            sum_col2.metric("Total Sum Assured", f"₹{total_sa:,.2f}")
            sum_col3.metric(
                "Total Premium (Incl. GST)", f"₹{total_prem_sum:,.2f}"
            )

            # --- DISPLAY GRID OUTPUT VIEW ---
            st.subheader("Output Preview")

            def highlight_calculated_cols(s):
                color = "background-color: yellow"
                targets = [
                    "Rate",
                    "Premium (Excl. GST)",
                    "GST amount",
                    "Total Premium (incl GST)",
                ]
                return [color if s.name in targets else "" for _ in s]

            st.dataframe(df_output.style.apply(highlight_calculated_cols, axis=0))

            # --- COMPILE FILE DOWN TO MEMORY BUFFER ---
            output_buffer = io.BytesIO()
            with pd.ExcelWriter(output_buffer, engine="openpyxl") as writer:
                df_output.to_excel(writer, index=False, sheet_name="GTL Premium")

                # Target spreadsheet formatting references
                worksheet = writer.sheets["GTL Premium"]
                yellow_fill = PatternFill(
                    start_color="FFFF00", end_color="FFFF00", fill_type="solid"
                )

                calc_col_indices = [
                    df_output.columns.get_loc("Rate") + 1,
                    df_output.columns.get_loc("Premium (Excl. GST)") + 1,
                    df_output.columns.get_loc("GST amount") + 1,
                    df_output.columns.get_loc("Total Premium (incl GST)") + 1,
                ]

                # Highlight calculation rows yellow without using dynamic cell formulas
                for col_idx in calc_col_indices:
                    for row_idx in range(2, len(df_output) + 2):
                        cell = worksheet.cell(row=row_idx, column=col_idx)
                        cell.fill = yellow_fill

                # Apply layout autofit directly across grid space
                for col in worksheet.columns:
                    max_len = max(len(str(cell.value or "")) for cell in col)
                    worksheet.column_dimensions[col[0].column_letter].width = (
                        max(max_len + 3, 12)
                    )

            excel_data = output_buffer.getvalue()

            st.download_button(
                label="Download Output Excel",
                data=excel_data,
                file_name="Bhumi_GTL_Premium_Output.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )

        except Exception as e:
            st.error(f"An error occurred while processing the file: {e}")
